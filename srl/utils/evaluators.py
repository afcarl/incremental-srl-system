import numpy as np

from misc import write
from vocab import UNDER_BAR


class Evaluator(object):
    def __init__(self, argv):
        self.argv = argv

    @staticmethod
    def _calc_f_measure(correct, r_total, p_total):
        precision = correct / p_total if p_total > 0 else 0.
        recall = correct / r_total if r_total > 0 else 0.
        f1 = (2 * precision * recall) / (precision + recall) if precision + recall > 0 else 0.
        return precision, recall, f1

    def f_measure_for_pi(self, y_true, y_pred):
        """
        :param y_true: 1D: n_batches, 2D: batch_size, 3D: n_words; elem=0/1
        :param y_pred: 1D: n_batches, 2D: batch_size, 3D: n_words; elem=0/1
        """
        p_total = 0.
        r_total = 0.
        correct = 0.

        for y_true_batch, y_pred_batch in zip(y_true, y_pred):
            y_true_batch = y_true_batch.flatten()
            y_pred_batch = y_pred_batch.flatten()

            eqs = np.equal(y_true_batch, y_pred_batch)
            correct += np.sum(eqs * np.greater(y_pred_batch, 0))
            r_total += np.sum(y_true_batch)
            p_total += np.sum(y_pred_batch)

        p, r, f = self._calc_f_measure(correct, r_total, p_total)
        write('\tF:{:>7.2%}  P:{:>7.2%} ({:>5}/{:>5})  R:{:>7.2%} ({:>5}/{:>5})\n'.format(
            f, p, int(correct), int(p_total), r, int(correct), int(r_total)))
        return f

    def f_measure_for_shift(self, y_true, y_pred):
        """
        :param y_true: 1D: n_batches, 2D: batch_size, 3D: n_words; elem=shift id
        :param y_pred: 1D: n_batches, 2D: batch_size, 3D: n_words; elem=shift id
        """
        p_total = 0.
        r_total = 0.
        correct = 0.

        for y_true_batch, y_pred_batch in zip(y_true, y_pred):
            y_true_batch = y_true_batch.flatten()
            y_pred_batch = y_pred_batch.flatten()

            eqs = np.equal(y_true_batch, y_pred_batch)
            correct += np.sum(eqs * np.greater(y_pred_batch, 0))
            r_total += np.sum(y_true_batch)
            p_total += np.sum(y_pred_batch)

        p, r, f = self._calc_f_measure(correct, r_total, p_total)
        write('\tF:{:>7.2%}  P:{:>7.2%} ({:>5}/{:>5})  R:{:>7.2%} ({:>5}/{:>5})\n'.format(
            f, p, int(correct), int(p_total), r, int(correct), int(r_total)))
        return f

    def f_measure_for_label(self, y_true, y_pred, vocab_label):
        """
        :param y_true: 1D: n_batches, 2D: batch_size, 3D: n_words; elem=label id
        :param y_pred: 1D: n_batches, 2D: batch_size, 3D: n_words; elem=label id
        :param vocab_label: Vocab()
        """
        p_total = 0.
        r_total = 0.
        correct = 0.
        none_id = vocab_label.get_id(UNDER_BAR)

        for y_true_batch, y_pred_batch in zip(y_true, y_pred):
            for y_true_i, y_pred_i in zip(y_true_batch, y_pred_batch):
                for y_true_j, y_pred_j in zip(y_true_i, y_pred_i):
                    if y_true_j != none_id:
                        r_total += 1
                    if y_pred_j != none_id:
                        p_total += 1
                    if y_true_j == y_pred_j and y_true_j != none_id:
                        correct += 1

        p, r, f = self._calc_f_measure(correct, r_total, p_total)
        write('\tF:{:>7.2%}  P:{:>7.2%} ({:>5}/{:>5})  R:{:>7.2%} ({:>5}/{:>5})\n'.format(
            f, p, int(correct), int(p_total), r, int(correct), int(r_total)))
        return f

    def isrl_label_score(self, label_true, label_pred, shift_ids, epoch=0):
        """
        :param label_true: 1D: n_batches, 2D: time_steps, 3D: batch_size * n_words(prd), 4D: n_words(arg); elem=label id
        :param label_pred: 1D: n_batches, 2D: time_steps, 3D: batch_size * n_words(prd) * n_words(arg); elem=label id
        :param shift_ids: 1D: n_batches, 2D: batch_size, 3D: n_words; elem=0/1
        :return: scalar
        """
        # 1D: n_calibrates, 2D: 3, 3D: n_words
        isrl_scores = np.zeros(shape=(4, 3, 40), dtype='float32')
        srl_score = np.zeros(shape=3, dtype='float32')

        for label_true_i, label_pred_i, shift_id_i in zip(label_true, label_pred, shift_ids):
            crr, true_total, pred_total = self._isrl_label_score(label_true_i, label_pred_i, shift_id_i)
            srl_score[0] += crr[-1]
            srl_score[1] += true_total[-1]
            srl_score[2] += pred_total[-1]

            n_words = len(label_true_i[0][0])
            if n_words <= 10:
                calib_id = 0
            elif n_words <= 20:
                calib_id = 1
            elif n_words <= 30:
                calib_id = 2
            elif n_words <= 40:
                calib_id = 3
            else:
                continue

            isrl_scores[calib_id, 0, :n_words] += crr
            isrl_scores[calib_id, 1, :n_words] += true_total
            isrl_scores[calib_id, 2, :n_words] += pred_total

        self._write_results(isrl_scores, srl_score, epoch)
        p, r, f = self._calc_f_measure(srl_score[0], srl_score[1], srl_score[2])
        return f

    @staticmethod
    def _isrl_label_score(label_true, label_pred, prd_marks):
        """
        :param label_true: 1D: time_steps, 2D: batch_size * n_words(prd), 3D: n_words(arg); elem=label id
        :param label_pred: 1D: time_steps, 2D: batch_size * n_words(prd) * n_words(arg); elem=label id
        :param prd_marks: 1D: batch_size, 2D: n_words; elem=0/1
        :return: 1D: time_steps
        """
        time_steps = label_pred.shape[0]
        batch_size = prd_marks.shape[0]
        n_words_a = prd_marks.shape[1]
        n_words_p = label_true.shape[1] / batch_size

        label_true = label_true.reshape((time_steps, batch_size, n_words_p, n_words_a))
        label_pred = label_pred.reshape((time_steps, batch_size, n_words_p, n_words_a))
        prd_marks = np.reshape(prd_marks, (1, batch_size, n_words_a, 1))

        eqs = np.equal(label_pred, label_true)
        crr = np.sum(eqs * prd_marks * np.greater(label_pred, 0), axis=(1, 2, 3))
        true_total = np.sum(np.greater(label_true, 0), axis=(1, 2, 3))
        pred_total = np.sum(np.greater(label_pred * prd_marks, 0), axis=(1, 2, 3))
        return crr, true_total, pred_total

    def _write_results(self, isrl_scores, srl_score, epoch):
        argv = self.argv
        if argv.output_fn:
            fn = 'result.' + argv.output_fn
        else:
            fn = 'result.layers-%s.e-%d.h-%d' % (argv.n_layers, argv.emb_dim, argv.hidden_dim)

        f = open(fn + '.txt', 'a')

        print >> f
        print >> f, '=' * 20
        print >> f, 'Epoch: %d\n' % epoch
        for calib_id, result in enumerate(isrl_scores):
            for i in xrange(40):
                if i >= (calib_id + 1) * 10:
                    break
                correct, r_total, p_total = result[:, i]
                p, r, f1 = self._calc_f_measure(correct, r_total, p_total)
                print >> f, '\tISRL-{} WORDS-{:>2} F:{:>7.2%} (P:{:>7.2%}|R:{:>7.2%}) ({:>5}|{:>5}|{:>5})'.format(
                    (calib_id + 1) * 10, i + 1, f1, p, r, int(correct), int(p_total), int(r_total))
            print >> f

        p, r, f1 = self._calc_f_measure(srl_score[0], srl_score[1], srl_score[2])
        print >> f, '\n\tSRL\tF:{:>7.2%}  P:{:>7.2%} ({:>5}/{:>5})  R:{:>7.2%} ({:>5}/{:>5})\n'.format(
            f1, p, int(srl_score[0]), int(srl_score[2]), r, int(srl_score[0]), int(srl_score[1]))

        f.close()

    def _write_results_xlsx(self, isrl_scores, srl_score):
        import openpyxl as px
        argv = self.argv
        if argv.output_fn:
            fn = 'result.' + argv.output_fn
        else:
            fn = 'result.layers-%s.e-%d.h-%d' % (argv.n_layers, argv.emb_dim, argv.hidden_dim)

        wb = px.Workbook()
        ws = wb.active

        ws.cell(row=1, column=2).value = 'F'
        ws.cell(row=1, column=3).value = 'P'
        ws.cell(row=1, column=4).value = 'R'
        ws.cell(row=1, column=5).value = '#CORRECT'
        ws.cell(row=1, column=6).value = '#P_TOTAL'
        ws.cell(row=1, column=7).value = '#R_TOTAL'
        row = 2
        for calib_id, result in enumerate(isrl_scores):
            for i in xrange(40):
                if i >= (calib_id + 1) * 10:
                    break
                correct, r_total, p_total = result[:, i]
                p, r, f1 = self._calc_f_measure(correct, r_total, p_total)

                ws.cell(row=row, column=2).value = f1
                ws.cell(row=row, column=3).value = p
                ws.cell(row=row, column=4).value = r
                ws.cell(row=row, column=5).value = int(correct)
                ws.cell(row=row, column=6).value = int(p_total)
                ws.cell(row=row, column=7).value = int(r_total)

                row += 1
            row += 1

        p, r, f1 = self._calc_f_measure(srl_score[0], srl_score[1], srl_score[2])
        ws.cell(row=row, column=2).value = f1
        ws.cell(row=row, column=3).value = p
        ws.cell(row=row, column=4).value = r
        ws.cell(row=row, column=5).value = int(srl_score[0])
        ws.cell(row=row, column=6).value = int(srl_score[2])
        ws.cell(row=row, column=7).value = int(srl_score[1])

        wb.save(fn + '.xlsx')


class ISRLEvaluator(Evaluator):
    def eval(self):
        argv = self.argv
        gold = self._load_file(argv.gold_data)
        pred = self._load_file(argv.pred_data)
        isrl_scores, srl_score = self.calc_score(gold, pred)
        if argv.output_type == 'txt':
            self._write_results(isrl_scores, srl_score, epoch=0)
        else:
            self._write_results_xlsx(isrl_scores, srl_score)

    @staticmethod
    def _load_file(fn):
        corpus = []
        sent = []
        prd = []
        boundary = ' ||| '
        with open(fn) as f:
            for line in f:
                if line.startswith('#'):
                    if prd:
                        sent.append(prd)
                        prd = []
                    if sent:
                        corpus.append(sent)
                    elem = [l for l in line[2:].rstrip().split(boundary)]
                    sent = [elem]
                elif line.startswith('--'):
                    elem = [l for l in line[3:].rstrip().split(boundary)]
                    prd.append(elem)
                elif line.startswith('-'):
                    if prd:
                        sent.append(prd)
                    elem = [l for l in line[2:].rstrip().split(boundary)]
                    prd = [elem]
            if prd:
                sent.append(prd)
                corpus.append(sent)

        return corpus

    def calc_score(self, gold, pred):
        # 1D: n_calibrates, 2D: 3, 3D: n_words
        isrl_scores = np.zeros(shape=(4, 3, 40), dtype='float32')
        srl_score = np.zeros(shape=3, dtype='float32')

        for i, sent_p in enumerate(pred):
            basic_info_p = sent_p[0]
            sent_index_p = int(basic_info_p[0])
            prd_indices_p = [int(p) for p in basic_info_p[1].split()]
            if prd_indices_p[0] == -1:
                prd_indices_p = []
            assert len(prd_indices_p) == len(sent_p[1:])

            sent_g = gold[i]
            basic_info_g = sent_g[0]
            sent_index_g = int(basic_info_g[0])
            prd_indices_g = [int(p) for p in basic_info_g[1].split()]
            if prd_indices_g[0] == -1:
                prd_indices_g = []
            assert len(prd_indices_g) == len(sent_g[1:])

            assert sent_index_g == sent_index_p

            srl_score += self._calc_srl_score(sent_g, sent_p, prd_indices_g, prd_indices_p)

            words = basic_info_p[-1].split()
            n_words = len(words)
            if n_words == 10:
                calib_id = 0
            elif n_words == 20:
                calib_id = 1
            elif n_words == 30:
                calib_id = 2
            elif n_words == 40:
                calib_id = 3
            else:
                continue

            isrl_scores[calib_id] += self._calc_isrl_score(sent_g, sent_p, prd_indices_g, prd_indices_p)
#            srl_score += self._calc_srl_score(sent_g, sent_p, prd_indices_g, prd_indices_p)

        return isrl_scores, srl_score

    @staticmethod
    def _calc_isrl_score(sent_g, sent_p, prd_indices_g, prd_indices_p):
        isrl_scores = np.zeros(shape=(3, 40), dtype='float32')
        for prd_i, prd_index_g in enumerate(prd_indices_g):
            prd_res_g = sent_g[prd_i + 1]
            if prd_index_g in prd_indices_p:
                prd_j = prd_indices_p.index(prd_index_g)
                prd_res_p = sent_p[prd_j + 1]
                assert len(prd_res_g) == len(prd_res_p)

                for time_step, (labels_g, labels_p) in enumerate(zip(prd_res_g[1:], prd_res_p[1:])):
                    labels_g = labels_g[1].split()
                    labels_p = labels_p[1].split()
                    assert len(labels_g) == len(labels_p)

                    for label_g, label_p in zip(labels_g, labels_p):
                        if label_g == label_p != '_':
                            isrl_scores[0][time_step] += 1

        for prd_res_g in sent_g[1:]:
            for time_step, labels_g in enumerate(prd_res_g[1:]):
                for label_g in labels_g[1].split():
                    if label_g != '_':
                        isrl_scores[1][time_step] += 1

        for prd_res_p in sent_p[1:]:
            for time_step, labels_p in enumerate(prd_res_p[1:]):
                for label_p in labels_p[1].split():
                    if label_p != '_':
                        isrl_scores[2][time_step] += 1

        return isrl_scores

    @staticmethod
    def _calc_srl_score(sent_g, sent_p, prd_indices_g, prd_indices_p):
        srl_scores = np.zeros(shape=3, dtype='float32')
        for prd_i, prd_index_g in enumerate(prd_indices_g):
            prd_res_g = sent_g[prd_i + 1]

            if prd_index_g in prd_indices_p:
                prd_j = prd_indices_p.index(prd_index_g)
                prd_res_p = sent_p[prd_j + 1]
                assert len(prd_res_g) == len(prd_res_p)

                labels_g = prd_res_g[0][-1].split()
                labels_p = prd_res_p[0][-1].split()
                assert len(labels_g) == len(labels_p)

                for label_g, label_p in zip(labels_g, labels_p):
                    if label_g == label_p != '_':
                        srl_scores[0] += 1

        for prd_res in sent_g[1:]:
            for label in prd_res[0][-1].split():
                if label != '_':
                    srl_scores[1] += 1

        for prd_res in sent_p[1:]:
            for label in prd_res[0][-1].split():
                if label != '_':
                    srl_scores[2] += 1

        return srl_scores

    def ciss(self, corpus):
        # 1D: n_calibrates, 2D: 3, 3D: n_words
        isrl_scores = np.zeros(shape=(4, 3, 40), dtype='float32')
        srl_score = np.zeros(shape=3, dtype='float32')

        for sent in corpus:
            n_words = sent.n_words
            if n_words <= 10:
                calib_id = 0
            elif n_words <= 20:
                calib_id = 1
            elif n_words <= 30:
                calib_id = 2
            elif n_words <= 40:
                calib_id = 3
            else:
                continue

            for i, p_index in enumerate(sent.prd_indices):
                labels_gold = sent.prd_props[i]
                for t in xrange(len(labels_gold)):
                    for label in labels_gold[:t+1]:
                        if label == UNDER_BAR:
                            continue
                        isrl_scores[calib_id, 1, t] += 1

                if p_index in sent.prd_indices_sys:
                    j = sent.prd_indices_sys.index(p_index)
                    labels_each_time = sent.prd_props_sys[j]
                    for t in xrange(len(labels_each_time)):
                        labels_gold_t = labels_gold[:t + 1]
                        labels_sys_t = labels_each_time[t]
                        for label_g, label_s in zip(labels_gold_t, labels_sys_t):
                            if label_g != UNDER_BAR and label_g == label_s:
                                isrl_scores[calib_id, 0, t] += 1

            for i, p_index in enumerate(sent.prd_indices_sys):
                for props in sent.prd_props_sys[i]:
                    for t, labels in enumerate(props):
                        for label in labels:
                            if label == UNDER_BAR:
                                continue
                            isrl_scores[calib_id, 2, t] += 1

        p, r, f = self._calc_f_measure(isrl_scores[0][0][-1], isrl_scores[0][1][-1], isrl_scores[0][2][-2])
        print f
        print p
        print r
        print isrl_scores
